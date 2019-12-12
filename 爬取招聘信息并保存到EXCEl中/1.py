#！/E:\python\env\python37\Scripts
# -*- coding:utf-8 -*-
#@Time  :2019/7/17 21:35
#@Author :zbwu103
#@File  ：1.py
#@功能：爬取51job中爬取的武汉的Python招聘信息，并且保存在excel中

import requests
from bs4 import BeautifulSoup
import openpyxl

def get_inf(page):
    url="https://search.51job.com/list/180200,000000,0000,00,9,99,python,2,{}.html".format(page)
    response=requests.get(url)
    response.encoding=response.apparent_encoding
    html=response.text
    bs=BeautifulSoup(html,'lxml')
    divs=bs.find('div',class_="dw_table").find_all("div",class_="el")
    inf={}
    for div in divs:
        if div.find("p",class_="t1"):
            inf[1]=div.find("p",class_="t1").find("a")['title']
            inf[2]=div.find("span",class_="t2").find("a")['title']
            inf[3]=div.find("span",class_="t3").text
            inf[4]=div.find("span",class_="t4").text
            inf[5]=div.find("span",class_="t5").text
            yield inf
def save_xl():
    wb=openpyxl.Workbook()
    sheet=wb["Sheet"]
    sheet.cell(row=1,column=1).value="职位"
    sheet.cell(row=1,column=2).value="公司名称"
    sheet.cell(row=1,column=3).value="地点"
    sheet.cell(row=1,column=4).value="月薪"
    sheet.cell(row=1,column=5).value="发布时间"
    r = 2
    for page in range(1,21):
        for inf in get_inf(page):
            for i in range(1,6):
                sheet.cell(row=r,column=i).value=inf[i]
            r += 1
    wb.save("招聘信息.xlsx")
def main():
    save_xl()

main()


